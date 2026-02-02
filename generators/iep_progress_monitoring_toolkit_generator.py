#!/usr/bin/env python3
"""
IEP Progress Monitoring Toolkit Generator

Generates a comprehensive toolkit for IEP data collection including:
- PDF print pack with quick-start guide and data collection sheets
- CSV/XLSX tracker templates for digital data management

Author: Small Wins Studio
License: Proprietary
"""

import json
import csv
import os
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor, black
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


class IEPProgressMonitoringGenerator:
    """Generator for IEP Progress Monitoring Toolkit materials."""
    
    def __init__(self, config_path: str):
        """
        Initialize the generator with a configuration file.
        
        Args:
            config_path: Path to JSON configuration file
        """
        with open(config_path, 'r') as f:
            self.config = json.load(f)
        
        # Page setup
        self.page_width, self.page_height = letter
        self.margins = self._get_margins()
        self.footer_text = f"© 2025 {self.config['brand']}. All rights reserved."
        
    def _get_margins(self) -> Dict[str, float]:
        """Get margins in points (1 inch = 72 points)."""
        margins = self.config['margins']
        return {
            'left': margins['left'] * inch,
            'right': margins['right'] * inch,
            'top': margins['top'] * inch,
            'bottom': margins['bottom'] * inch
        }
    
    def _draw_footer(self, c: canvas.Canvas, page_num: Optional[int] = None):
        """Draw footer on the page."""
        c.saveState()
        c.setFont(self.config['fonts']['footer']['family'], 
                  self.config['fonts']['footer']['size'])
        c.setFillColor(HexColor(self.config['colors']['footer_text']))
        
        footer_y = self.margins['bottom'] / 2
        c.drawString(self.margins['left'], footer_y, self.footer_text)
        
        if page_num:
            c.drawRightString(
                self.page_width - self.margins['right'],
                footer_y,
                f"Page {page_num}"
            )
        
        c.restoreState()
    
    def _draw_page_title(self, c: canvas.Canvas, title: str, subtitle: str = ""):
        """Draw page title and subtitle."""
        y_position = self.page_height - self.margins['top'] - 30
        
        # Draw title
        c.setFont(self.config['fonts']['title']['family'],
                  self.config['fonts']['title']['size'])
        c.setFillColor(black)
        c.drawString(self.margins['left'], y_position, title)
        
        # Draw subtitle if provided
        if subtitle:
            y_position -= 25
            c.setFont(self.config['fonts']['subtitle']['family'],
                      self.config['fonts']['subtitle']['size'])
            c.setFillColor(HexColor(self.config['colors']['footer_text']))
            c.drawString(self.margins['left'], y_position, subtitle)
        
        return y_position - 20  # Return next available y position
    
    def generate_quick_start_pages(self, c: canvas.Canvas) -> int:
        """
        Generate quick-start guide pages.
        
        Returns:
            Number of pages generated
        """
        page_count = 0
        
        for page_config in self.config['quick_start_pages']:
            c.showPage()
            page_count += 1
            
            # Draw title
            y_pos = self._draw_page_title(c, page_config['title'], 
                                         page_config.get('subtitle', ''))
            y_pos -= 20
            
            # Handle table type page
            if page_config.get('type') == 'table':
                self._draw_reference_table(c, y_pos, page_config['table_data'])
            else:
                # Handle sections with content
                for section in page_config.get('sections', []):
                    y_pos = self._draw_section(c, y_pos, section)
            
            # Draw footer
            self._draw_footer(c, page_count)
        
        return page_count
    
    def _draw_section(self, c: canvas.Canvas, y_pos: float, section: Dict) -> float:
        """Draw a content section and return new y position."""
        # Draw heading
        c.setFont(self.config['fonts']['subtitle']['family'],
                  self.config['fonts']['subtitle']['size'])
        c.setFillColor(black)
        c.drawString(self.margins['left'], y_pos, section['heading'])
        y_pos -= 20
        
        # Draw content lines
        c.setFont(self.config['fonts']['body']['family'],
                  self.config['fonts']['body']['size'])
        
        for line in section['content']:
            if y_pos < self.margins['bottom'] + 50:
                break  # Prevent footer overlap
            c.drawString(self.margins['left'], y_pos, line)
            y_pos -= 15
        
        y_pos -= 10  # Extra space after section
        return y_pos
    
    def _draw_reference_table(self, c: canvas.Canvas, y_pos: float, table_data: Dict):
        """Draw a reference table using reportlab Table."""
        data = [table_data['headers']] + table_data['rows']
        
        # Calculate column widths
        usable_width = self.page_width - self.margins['left'] - self.margins['right']
        col_widths = [usable_width * 0.3, usable_width * 0.3, usable_width * 0.4]
        
        # Create table
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), HexColor(self.config['colors']['table_header_bg'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), black),
            ('FONTNAME', (0, 0), (-1, 0), self.config['fonts']['table_header']['family']),
            ('FONTSIZE', (0, 0), (-1, 0), self.config['fonts']['table_header']['size']),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Body styling
            ('FONTNAME', (0, 1), (-1, -1), self.config['fonts']['table_body']['family']),
            ('FONTSIZE', (0, 1), (-1, -1), self.config['fonts']['table_body']['size']),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 1, HexColor(self.config['colors']['table_border'])),
        ]))
        
        # Draw table
        table_width, table_height = table.wrap(usable_width, self.page_height)
        table.drawOn(c, self.margins['left'], y_pos - table_height)
    
    def generate_data_sheets(self, c: canvas.Canvas, start_page: int) -> int:
        """
        Generate all data collection sheet templates.
        
        Args:
            c: Canvas object
            start_page: Starting page number
            
        Returns:
            Number of pages generated
        """
        page_count = start_page
        
        for sheet_config in self.config['data_sheets']:
            c.showPage()
            page_count += 1
            
            sheet_type = sheet_config['type']
            
            if sheet_type == 'trials_accuracy':
                self._generate_trials_accuracy_sheet(c, sheet_config)
            elif sheet_type == 'frequency_count':
                self._generate_frequency_count_sheet(c, sheet_config)
            elif sheet_type == 'duration':
                self._generate_duration_sheet(c, sheet_config)
            elif sheet_type in ['interval_5min', 'interval_10min']:
                self._generate_interval_sheet(c, sheet_config)
            elif sheet_type == 'abc_observation':
                self._generate_abc_observation_sheet(c, sheet_config)
            elif sheet_type == 'work_sample':
                self._generate_work_sample_sheet(c, sheet_config)
            elif sheet_type == 'goal_snapshot':
                self._generate_goal_snapshot_sheet(c, sheet_config)
            
            self._draw_footer(c, page_count)
        
        return page_count - start_page
    
    def _draw_field_line(self, c: canvas.Canvas, x: float, y: float, 
                         label: str, width: float = 200):
        """Draw a labeled field with a line for writing."""
        c.setFont(self.config['fonts']['body']['family'],
                  self.config['fonts']['body']['size'])
        c.drawString(x, y, label)
        
        label_width = c.stringWidth(label, 
                                    self.config['fonts']['body']['family'],
                                    self.config['fonts']['body']['size'])
        
        # Draw line
        line_x = x + label_width + 5
        c.line(line_x, y - 2, line_x + width, y - 2)
    
    def _generate_trials_accuracy_sheet(self, c: canvas.Canvas, config: Dict):
        """Generate Trials/Accuracy (DTT-style) data sheet."""
        y_pos = self._draw_page_title(c, config['title'])
        y_pos -= 30
        
        # Student info fields
        self._draw_field_line(c, self.margins['left'], y_pos, 
                             "Student Name:", 250)
        self._draw_field_line(c, self.page_width / 2, y_pos, 
                             "Date:", 150)
        y_pos -= 25
        
        self._draw_field_line(c, self.margins['left'], y_pos, 
                             "Goal/Skill:", 400)
        y_pos -= 35
        
        # Prompt level legend
        c.setFont(self.config['fonts']['body']['family'], 8)
        legend_text = "Prompt Levels: " + " | ".join([
            f"{k}={v}" for k, v in config['prompt_legend'].items()
        ])
        c.drawString(self.margins['left'], y_pos, legend_text)
        y_pos -= 25
        
        # Create data table
        headers = ['Trial']
        for i in range(1, config['sessions'] + 1):
            headers.extend([f'Session {i}', 'Prompt'])
        
        # Build table data
        table_data = [headers]
        for trial in range(1, config['trials'] + 1):
            row = [str(trial)]
            for _ in range(config['sessions']):
                row.extend(['', ''])  # Response and Prompt columns
            table_data.append(row)
        
        # Add accuracy row
        accuracy_row = ['% Correct']
        for _ in range(config['sessions']):
            accuracy_row.extend(['', ''])
        table_data.append(accuracy_row)
        
        # Calculate column widths
        usable_width = self.page_width - self.margins['left'] - self.margins['right']
        col_widths = [40] + [50, 35] * config['sessions']
        
        # Adjust if too wide
        if sum(col_widths) > usable_width:
            scale = usable_width / sum(col_widths)
            col_widths = [w * scale for w in col_widths]
        
        # Create and style table
        table = Table(table_data, colWidths=col_widths, rowHeights=20)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), self.config['fonts']['table_header']['family']),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('FONTNAME', (0, 1), (-1, -1), self.config['fonts']['table_body']['family']),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BACKGROUND', (0, 0), (-1, 0), HexColor(self.config['colors']['table_header_bg'])),
            ('BACKGROUND', (0, -1), (-1, -1), HexColor('#F5F5F5')),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor(self.config['colors']['table_border'])),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        table_width, table_height = table.wrap(usable_width, self.page_height)
        table.drawOn(c, self.margins['left'], y_pos - table_height)
        y_pos -= table_height + 20
        
        # Notes section
        if y_pos > self.margins['bottom'] + 50:
            c.setFont(self.config['fonts']['body']['family'], 
                     self.config['fonts']['body']['size'])
            c.drawString(self.margins['left'], y_pos, "Notes:")
            y_pos -= 5
            for i in range(3):
                if y_pos < self.margins['bottom'] + 50:
                    break
                c.line(self.margins['left'], y_pos, 
                      self.page_width - self.margins['right'], y_pos)
                y_pos -= 15
    
    def _generate_frequency_count_sheet(self, c: canvas.Canvas, config: Dict):
        """Generate Frequency Count data sheet."""
        y_pos = self._draw_page_title(c, config['title'])
        y_pos -= 30
        
        # Student info
        self._draw_field_line(c, self.margins['left'], y_pos, 
                             "Student Name:", 250)
        self._draw_field_line(c, self.page_width / 2, y_pos, 
                             "Week of:", 150)
        y_pos -= 25
        
        self._draw_field_line(c, self.margins['left'], y_pos, 
                             "Target Behavior:", 400)
        y_pos -= 35
        
        # Create table
        table_data = [['Date', 'Time Period', 'Tally Marks', 'Total Count', 'Notes']]
        
        for _ in range(config['sessions']):
            table_data.append(['', '', '', '', ''])
        
        # Column widths
        usable_width = self.page_width - self.margins['left'] - self.margins['right']
        col_widths = [70, 80, 120, 70, usable_width - 340]
        
        table = Table(table_data, colWidths=col_widths, rowHeights=25)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), self.config['fonts']['table_header']['family']),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTNAME', (0, 1), (-1, -1), self.config['fonts']['table_body']['family']),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), HexColor(self.config['colors']['table_header_bg'])),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor(self.config['colors']['table_border'])),
            ('ALIGN', (0, 0), (3, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        table_width, table_height = table.wrap(usable_width, self.page_height)
        table.drawOn(c, self.margins['left'], y_pos - table_height)
    
    def _generate_duration_sheet(self, c: canvas.Canvas, config: Dict):
        """Generate Duration Tracking data sheet."""
        y_pos = self._draw_page_title(c, config['title'])
        y_pos -= 30
        
        # Student info
        self._draw_field_line(c, self.margins['left'], y_pos, 
                             "Student Name:", 250)
        self._draw_field_line(c, self.page_width / 2, y_pos, 
                             "Week of:", 150)
        y_pos -= 25
        
        self._draw_field_line(c, self.margins['left'], y_pos, 
                             "Target Behavior/Activity:", 400)
        y_pos -= 35
        
        # Create table
        table_data = [['Date', 'Start Time', 'End Time', 'Total Duration', 'Notes']]
        
        for _ in range(config['sessions']):
            table_data.append(['', '', '', '', ''])
        
        # Add summary row
        table_data.append(['', '', 'Average:', '', ''])
        
        # Column widths
        usable_width = self.page_width - self.margins['left'] - self.margins['right']
        col_widths = [70, 70, 70, 90, usable_width - 300]
        
        table = Table(table_data, colWidths=col_widths, rowHeights=25)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), self.config['fonts']['table_header']['family']),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTNAME', (0, 1), (-1, -1), self.config['fonts']['table_body']['family']),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), HexColor(self.config['colors']['table_header_bg'])),
            ('BACKGROUND', (0, -1), (-1, -1), HexColor('#F5F5F5')),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor(self.config['colors']['table_border'])),
            ('ALIGN', (0, 0), (3, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        table_width, table_height = table.wrap(usable_width, self.page_height)
        table.drawOn(c, self.margins['left'], y_pos - table_height)
    
    def _generate_interval_sheet(self, c: canvas.Canvas, config: Dict):
        """Generate Interval Recording data sheet."""
        y_pos = self._draw_page_title(c, config['title'])
        y_pos -= 30
        
        # Student info
        self._draw_field_line(c, self.margins['left'], y_pos, 
                             "Student Name:", 250)
        self._draw_field_line(c, self.page_width / 2, y_pos, 
                             "Date:", 150)
        y_pos -= 25
        
        self._draw_field_line(c, self.margins['left'], y_pos, 
                             "Target Behavior:", 400)
        y_pos -= 30
        
        # Instructions
        c.setFont(self.config['fonts']['body']['family'], 8)
        c.drawString(self.margins['left'], y_pos, 
                    f"Mark '+' if behavior occurs during interval, '−' if not. "
                    f"Session length: {config['session_length']} minutes")
        y_pos -= 25
        
        # Create interval table
        interval_dur = config['interval_duration']
        headers = ['Session/Date']
        for i in range(config['intervals']):
            start = i * interval_dur
            end = (i + 1) * interval_dur
            headers.append(f'{start}-{end} min')
        headers.append('% Intervals')
        
        table_data = [headers]
        
        # Add 5 session rows
        for i in range(1, 6):
            row = [f'Session {i}']
            row.extend([''] * config['intervals'])
            row.append('')
            table_data.append(row)
        
        # Column widths
        usable_width = self.page_width - self.margins['left'] - self.margins['right']
        col_widths = [80] + [60] * config['intervals'] + [70]
        
        if sum(col_widths) > usable_width:
            scale = usable_width / sum(col_widths)
            col_widths = [w * scale for w in col_widths]
        
        table = Table(table_data, colWidths=col_widths, rowHeights=30)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), self.config['fonts']['table_header']['family']),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('FONTNAME', (0, 1), (-1, -1), self.config['fonts']['table_body']['family']),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), HexColor(self.config['colors']['table_header_bg'])),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor(self.config['colors']['table_border'])),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        table_width, table_height = table.wrap(usable_width, self.page_height)
        table.drawOn(c, self.margins['left'], y_pos - table_height)
        y_pos -= table_height + 20
        
        # Notes
        if y_pos > self.margins['bottom'] + 50:
            c.setFont(self.config['fonts']['body']['family'], 
                     self.config['fonts']['body']['size'])
            c.drawString(self.margins['left'], y_pos, "Notes:")
            y_pos -= 5
            for i in range(2):
                if y_pos < self.margins['bottom'] + 50:
                    break
                c.line(self.margins['left'], y_pos, 
                      self.page_width - self.margins['right'], y_pos)
                y_pos -= 15
    
    def _generate_abc_observation_sheet(self, c: canvas.Canvas, config: Dict):
        """Generate ABC Observation Form."""
        y_pos = self._draw_page_title(c, config['title'])
        y_pos -= 30
        
        # Student info
        self._draw_field_line(c, self.margins['left'], y_pos, 
                             "Student Name:", 250)
        self._draw_field_line(c, self.page_width / 2, y_pos, 
                             "Observer:", 150)
        y_pos -= 25
        
        self._draw_field_line(c, self.margins['left'], y_pos, 
                             "Target Behavior:", 400)
        y_pos -= 35
        
        # Create ABC table
        table_data = [['Date/Time', 'Antecedent\n(What happened before?)', 
                      'Behavior\n(What did student do?)', 
                      'Consequence\n(What happened after?)']]
        
        for _ in range(config['rows']):
            table_data.append(['', '', '', ''])
        
        # Column widths
        usable_width = self.page_width - self.margins['left'] - self.margins['right']
        col_widths = [80, usable_width * 0.3, usable_width * 0.3, usable_width * 0.3]
        actual_total = 80 + usable_width * 0.9
        if actual_total > usable_width:
            col_widths = [80, (usable_width - 80) / 3, 
                         (usable_width - 80) / 3, (usable_width - 80) / 3]
        
        table = Table(table_data, colWidths=col_widths, rowHeights=35)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), self.config['fonts']['table_header']['family']),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTNAME', (0, 1), (-1, -1), self.config['fonts']['table_body']['family']),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), HexColor(self.config['colors']['table_header_bg'])),
            ('GRID', (0, 0), (-1, -1), 0.5, HexColor(self.config['colors']['table_border'])),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        table_width, table_height = table.wrap(usable_width, self.page_height)
        table.drawOn(c, self.margins['left'], y_pos - table_height)
    
    def _generate_work_sample_sheet(self, c: canvas.Canvas, config: Dict):
        """Generate Work Sample / Anecdotal Log."""
        y_pos = self._draw_page_title(c, config['title'])
        y_pos -= 30
        
        # Student info
        self._draw_field_line(c, self.margins['left'], y_pos, 
                             "Student Name:", 250)
        self._draw_field_line(c, self.page_width / 2, y_pos, 
                             "Date:", 150)
        y_pos -= 25
        
        self._draw_field_line(c, self.margins['left'], y_pos, 
                             "Activity/Task:", 400)
        y_pos -= 35
        
        # Observation section
        c.setFont(self.config['fonts']['subtitle']['family'], 11)
        c.drawString(self.margins['left'], y_pos, "Observation Notes:")
        y_pos -= 20
        
        c.setFont(self.config['fonts']['body']['family'], 
                 self.config['fonts']['body']['size'])
        
        # Draw lines for writing
        available_height = y_pos - self.margins['bottom'] - 100
        line_spacing = 18
        num_lines = int(available_height / line_spacing)
        
        for i in range(num_lines):
            if y_pos < self.margins['bottom'] + 100:
                break
            c.line(self.margins['left'], y_pos, 
                  self.page_width - self.margins['right'], y_pos)
            y_pos -= line_spacing
        
        # Next steps section
        y_pos = self.margins['bottom'] + 80
        c.setFont(self.config['fonts']['subtitle']['family'], 11)
        c.drawString(self.margins['left'], y_pos, "Next Steps:")
        y_pos -= 20
        
        c.setFont(self.config['fonts']['body']['family'], 
                 self.config['fonts']['body']['size'])
        for i in range(3):
            c.line(self.margins['left'], y_pos, 
                  self.page_width - self.margins['right'], y_pos)
            y_pos -= 18
    
    def _generate_goal_snapshot_sheet(self, c: canvas.Canvas, config: Dict):
        """Generate Goal Snapshot Page."""
        y_pos = self._draw_page_title(c, config['title'])
        y_pos -= 30
        
        # Student info
        self._draw_field_line(c, self.margins['left'], y_pos, 
                             "Student Name:", 250)
        self._draw_field_line(c, self.page_width / 2, y_pos, 
                             "Review Date:", 150)
        y_pos -= 35
        
        # IEP Goal section
        c.setFont(self.config['fonts']['subtitle']['family'], 11)
        c.drawString(self.margins['left'], y_pos, "IEP Goal Statement:")
        y_pos -= 20
        
        c.setFont(self.config['fonts']['body']['family'], 
                 self.config['fonts']['body']['size'])
        for i in range(3):
            c.line(self.margins['left'], y_pos, 
                  self.page_width - self.margins['right'], y_pos)
            y_pos -= 18
        
        y_pos -= 10
        
        # Current Performance
        c.setFont(self.config['fonts']['subtitle']['family'], 11)
        c.drawString(self.margins['left'], y_pos, "Current Level of Performance:")
        y_pos -= 20
        
        c.setFont(self.config['fonts']['body']['family'], 
                 self.config['fonts']['body']['size'])
        for i in range(2):
            c.line(self.margins['left'], y_pos, 
                  self.page_width - self.margins['right'], y_pos)
            y_pos -= 18
        
        y_pos -= 10
        
        # Supports/Accommodations
        c.setFont(self.config['fonts']['subtitle']['family'], 11)
        c.drawString(self.margins['left'], y_pos, "Supports/Accommodations:")
        y_pos -= 20
        
        c.setFont(self.config['fonts']['body']['family'], 
                 self.config['fonts']['body']['size'])
        for i in range(3):
            c.line(self.margins['left'], y_pos, 
                  self.page_width - self.margins['right'], y_pos)
            y_pos -= 18
        
        y_pos -= 10
        
        # Success Criteria
        c.setFont(self.config['fonts']['subtitle']['family'], 11)
        c.drawString(self.margins['left'], y_pos, "Success Criteria:")
        y_pos -= 20
        
        c.setFont(self.config['fonts']['body']['family'], 
                 self.config['fonts']['body']['size'])
        for i in range(2):
            c.line(self.margins['left'], y_pos, 
                  self.page_width - self.margins['right'], y_pos)
            y_pos -= 18
        
        y_pos -= 10
        
        # Progress Notes
        c.setFont(self.config['fonts']['subtitle']['family'], 11)
        c.drawString(self.margins['left'], y_pos, "Progress Notes:")
        y_pos -= 20
        
        c.setFont(self.config['fonts']['body']['family'], 
                 self.config['fonts']['body']['size'])
        
        available_lines = int((y_pos - self.margins['bottom'] - 50) / 18)
        for i in range(min(available_lines, 8)):
            c.line(self.margins['left'], y_pos, 
                  self.page_width - self.margins['right'], y_pos)
            y_pos -= 18
    
    def generate_pdf(self, output_path: str):
        """
        Generate the complete PDF toolkit.
        
        Args:
            output_path: Path where the PDF should be saved
        """
        c = canvas.Canvas(output_path, pagesize=letter)
        
        # Generate quick-start pages
        quick_start_pages = self.generate_quick_start_pages(c)
        
        # Generate data sheets
        data_sheet_pages = self.generate_data_sheets(c, quick_start_pages)
        
        c.save()
        
        print(f"✓ Generated PDF: {output_path}")
        print(f"  - Quick-start pages: {quick_start_pages}")
        print(f"  - Data sheet pages: {data_sheet_pages}")
        print(f"  - Total pages: {quick_start_pages + data_sheet_pages}")
    
    def generate_csv_tracker(self, output_dir: str):
        """
        Generate CSV tracker templates.
        
        Args:
            output_dir: Directory where CSV files should be saved
        """
        tracker_config = self.config['tracker_template']
        
        for tab_config in tracker_config['tabs']:
            filename = f"{tab_config['name']}.csv"
            filepath = os.path.join(output_dir, filename)
            
            with open(filepath, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write header
                writer.writerow(tab_config['columns'])
                
                # Write sample rows (empty)
                for _ in range(5):
                    writer.writerow([''] * len(tab_config['columns']))
            
            print(f"✓ Generated CSV: {filepath}")
    
    def generate_xlsx_tracker(self, output_path: str):
        """
        Generate XLSX tracker template with formatting.
        
        Args:
            output_path: Path where the XLSX file should be saved
        """
        if not XLSX_AVAILABLE:
            print("⚠ XLSX generation skipped (openpyxl not available)")
            return
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        tracker_config = self.config['tracker_template']
        
        # Define styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", 
                                   fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        
        for tab_config in tracker_config['tabs']:
            ws = wb.create_sheet(title=tab_config['title'])
            
            # Write and format headers
            for col_idx, header in enumerate(tab_config['columns'], start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
            
            # Add empty rows
            for row_idx in range(2, 12):
                for col_idx in range(1, len(tab_config['columns']) + 1):
                    ws.cell(row=row_idx, column=col_idx, value="")
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        print(f"✓ Generated XLSX: {output_path}")


def main():
    """Main execution function."""
    # Paths
    script_dir = Path(__file__).parent.parent
    config_path = script_dir / "data" / "iep_progress_monitoring" / "sample_config.json"
    output_dir = script_dir / "samples" / "iep_progress_monitoring"
    
    # Ensure output directory exists
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Initialize generator
    print("IEP Progress Monitoring Toolkit Generator")
    print("=" * 50)
    print(f"Config: {config_path}")
    print(f"Output: {output_dir}")
    print()
    
    generator = IEPProgressMonitoringGenerator(str(config_path))
    
    # Generate PDF
    pdf_path = output_dir / "iep_progress_monitoring_toolkit_sample.pdf"
    generator.generate_pdf(str(pdf_path))
    print()
    
    # Generate CSV tracker
    generator.generate_csv_tracker(str(output_dir))
    print()
    
    # Generate XLSX tracker
    xlsx_path = output_dir / "iep_tracker_sample.xlsx"
    generator.generate_xlsx_tracker(str(xlsx_path))
    print()
    
    print("=" * 50)
    print("✓ Toolkit generation complete!")
    print(f"\nGenerated files in: {output_dir}")


if __name__ == "__main__":
    main()
